import os
from datetime import date, datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "Blinkit PO Tracker.xlsx")

# Maps our extracted keys → Excel column headers (exact header text in row 1)
COLUMN_MAP = {
    "CHAINS": "CHAINS",
    "SITE CODE": "SITE CODE ",       # note trailing space in original header
    "VENDOR CODE": "VENDOR CODE",
    "VENDOR NAME": "VENDOR NAME",
    "PO NO": "PO NO",
    "PO DATE": "PO DATE",
    "DELIVERY DATE": "DELIVERY DATE",
    "ARTICLE DESCRIPTION": "ARTICLE DESCRIPTION",
    "TOTAL PCS": "TOTAL PCS",
    "LANDING PRICE": "LANDING PRICE",
    "TOTAL BASIC PO VALUE WITH TAX": "TOTAL BASIC PO VALUE WITH TAX",
}


def _load_workbook():
    path = os.path.abspath(EXCEL_PATH)
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel file not found: {path}")
    return openpyxl.load_workbook(path), path


def _get_header_map(ws) -> dict[str, int]:
    """Return {header_text: col_index (1-based)} from row 1."""
    header_map = {}
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value is not None:
                header_map[str(cell.value).strip()] = cell.column
    return header_map


def append_rows(rows: list[dict]) -> dict:
    """Append extracted PO rows to Sheet1 and save."""
    wb, path = _load_workbook()
    ws = wb["Sheet1"]

    # Build header→col map (strip spaces for matching)
    raw_headers = {}
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value is not None:
                raw_headers[str(cell.value).strip()] = cell.column

    # Find the next empty row
    next_row = ws.max_row + 1
    # Check if last row is actually empty (openpyxl can report wrong max_row)
    while next_row > 2:
        row_vals = [ws.cell(row=next_row - 1, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None for v in row_vals):
            break
        next_row -= 1

    written = 0
    for row_data in rows:
        for key, header in COLUMN_MAP.items():
            # Match header ignoring trailing spaces
            col_idx = None
            for h, ci in raw_headers.items():
                if h.upper() == header.strip().upper():
                    col_idx = ci
                    break
            if col_idx is None:
                continue

            value = row_data.get(key)
            cell = ws.cell(row=next_row, column=col_idx, value=value)

            # Format dates
            if isinstance(value, (date, datetime)):
                cell.number_format = "DD-MMM-YYYY"
        next_row += 1
        written += 1

    wb.save(path)
    return {"appended": written, "path": path}


def clear_sheet() -> dict:
    """Delete all data rows from Sheet1 (keep header row 1)."""
    wb, path = _load_workbook()
    ws = wb["Sheet1"]

    rows_deleted = 0
    # Delete from bottom up to avoid index shifting
    for row in range(ws.max_row, 1, -1):
        row_vals = [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None for v in row_vals):
            ws.delete_rows(row)
            rows_deleted += 1

    wb.save(path)
    return {"deleted": rows_deleted, "path": path}


def get_sheet_preview(max_rows: int = 50) -> dict:
    """Return headers and first N data rows as JSON-friendly dicts."""
    wb, path = _load_workbook()
    ws = wb["Sheet1"]

    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value) if cell.value is not None else "")

    data_rows = []
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, max_rows + 1), values_only=True):
        if any(v is not None for v in row):
            serialized = []
            for v in row:
                if isinstance(v, (date, datetime)):
                    serialized.append(v.isoformat())
                else:
                    serialized.append(v)
            data_rows.append(serialized)

    return {"headers": headers, "rows": data_rows, "total_rows": len(data_rows)}
